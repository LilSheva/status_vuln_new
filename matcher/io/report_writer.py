"""XLSX report generation with 3 sheets."""

from __future__ import annotations

import logging
from datetime import datetime, timedelta
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from shared.constants import (
    STATUS_CONDITIONAL,
    STATUS_LINUX,
    STATUS_NO,
    STATUS_REPEAT,
    STATUS_YES,
)

if TYPE_CHECKING:
    from pathlib import Path

    from openpyxl.worksheet.worksheet import Worksheet

    from shared.types import AnalysisResult, JournalEntry, MatchCandidate, PipelineSettings

logger = logging.getLogger(__name__)

_HEADER_FONT = Font(name="Segoe UI", bold=True, size=11, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="4A5ABF", end_color="4A5ABF", fill_type="solid")
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
_THIN_SIDE = Side(style="thin", color="000000")
_SEP_SIDE = Side(style="medium", color="000000")   # group separator in detail sheet

_THIN_BORDER = Border(
    left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE,
)

_STATUS_FONT_COLORS: dict[str, str] = {
    STATUS_YES: "C0392B",       # red
    STATUS_NO: "27AE60",        # green
    STATUS_LINUX: "2980B9",     # blue
    STATUS_CONDITIONAL: "E67E22",  # orange
    STATUS_REPEAT: "C0392B",    # red
}

_SCORE_FILLS = {
    "high": PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid"),
    "medium": PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid"),
    "low": PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid"),
}


def _style_header(ws: Worksheet, col_count: int) -> None:
    """Apply header styles to the first row."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT
        cell.border = _THIN_BORDER


def _auto_width(ws: Worksheet, col_count: int, max_width: int = 50) -> None:
    """Auto-fit column widths based on content."""
    for col in range(1, col_count + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        adjusted = min(max_len + 4, max_width)
        ws.column_dimensions[get_column_letter(col)].width = max(adjusted, 10)


def _apply_table_borders(ws: Worksheet, last_row: int, col_count: int) -> None:
    """Apply black thin borders to every cell of the table (rows 1..last_row)."""
    for row in range(1, last_row + 1):
        for col in range(1, col_count + 1):
            ws.cell(row=row, column=col).border = _THIN_BORDER


def _score_fill(score: float) -> PatternFill | None:
    """Return a color fill based on score value."""
    if score >= 0.7:
        return _SCORE_FILLS["high"]
    if score >= 0.4:
        return _SCORE_FILLS["medium"]
    if score > 0:
        return _SCORE_FILLS["low"]
    return None


def _candidate_tier(score: float) -> int:
    """Return score tier: 1=high (≥0.7), 2=medium (≥0.4), 3=low (<0.4)."""
    if score >= 0.7:
        return 1
    if score >= 0.4:
        return 2
    return 3


def _filter_candidates(
    candidates: list[MatchCandidate],
    primary_limit: int = 0,
    secondary_limit: int = 3,
) -> list[MatchCandidate]:
    """Filter candidates by tier priority (best tier + secondary_limit from next tier)."""
    tier1 = [c for c in candidates if _candidate_tier(c.combined_score) == 1]
    tier2 = [c for c in candidates if _candidate_tier(c.combined_score) == 2]
    tier3 = [c for c in candidates if _candidate_tier(c.combined_score) == 3]

    if tier1:
        primary = tier1 if primary_limit == 0 else tier1[:primary_limit]
        return primary + tier2[:secondary_limit]
    if tier2:
        primary = tier2 if primary_limit == 0 else tier2[:primary_limit]
        return primary + tier3[:secondary_limit]
    return tier3


def _vendor_product(vendor: str, name: str) -> str:
    """Return 'vendor - product' or just product if vendor is empty."""
    return f"{vendor} - {name}" if vendor else name


# ---------------------------------------------------------------------------
# Sheet 1: Main table
# ---------------------------------------------------------------------------

_MAIN_HEADERS = [
    "№",
    "Дата",
    "Ответственный",
    "Публикация",
    "Статус",
    "ID ППТС",
    "CVE",
    "CVSS",
    "Продукт",
    "Источник",
]


def _report_date() -> str:
    """Return report date (yesterday if before 08:00, today otherwise)."""
    now = datetime.now()
    date = (now - timedelta(days=1)).date() if now.hour < 8 else now.date()
    return date.strftime("%d.%m.%Y")


def _write_main_sheet(
    ws: Worksheet,
    results: list[AnalysisResult],
    responsible: str,
    publication: str,
) -> None:
    """Write the main results table (Sheet 1)."""
    ws.title = "Основная таблица"

    for col, header in enumerate(_MAIN_HEADERS, 1):
        ws.cell(row=1, column=col, value=header)
    _style_header(ws, len(_MAIN_HEADERS))

    date_str = _report_date()

    for row_idx, r in enumerate(results, 2):
        v = r.vulnerability
        num = row_idx - 1

        if r.status == STATUS_YES:
            ppts_value = r.ppts_id or ""
        elif r.status in (STATUS_NO, STATUS_CONDITIONAL, STATUS_LINUX):
            ppts_value = "----------"
        else:
            ppts_value = ""

        ws.cell(row=row_idx, column=1, value=num)
        ws.cell(row=row_idx, column=2, value=date_str)
        ws.cell(row=row_idx, column=3, value=responsible)
        ws.cell(row=row_idx, column=4, value=publication)

        status_cell = ws.cell(row=row_idx, column=5, value=r.status or None)
        if r.status in _STATUS_FONT_COLORS:
            status_cell.font = Font(color=_STATUS_FONT_COLORS[r.status], bold=True)
        status_cell.alignment = Alignment(horizontal="center")

        ws.cell(row=row_idx, column=6, value=ppts_value)
        ws.cell(row=row_idx, column=7, value=v.cve_id)
        ws.cell(row=row_idx, column=8, value=v.cvss or "")
        ws.cell(row=row_idx, column=9, value=_vendor_product(v.vendor, v.product))
        ws.cell(row=row_idx, column=10, value=v.source_url or "")

    last_data_row = 1 + len(results)
    _apply_table_borders(ws, last_data_row, len(_MAIN_HEADERS))

    _auto_width(ws, len(_MAIN_HEADERS))
    ws.auto_filter.ref = ws.dimensions


# ---------------------------------------------------------------------------
# Sheet 2: Detailed analysis
# ---------------------------------------------------------------------------

_DETAIL_HEADERS = [
    "№",
    "CVE ID",
    "Продукт (ТСУ)",
    "Кандидат (ППТС)",
    "ППТС ID",
    "Vector Score",
    "Fuzzy Score",
    "Exact Score",
    "Итоговый балл",
    "Ранг",
    "Источник",
    "Журнал (файл)",
    "Журнал (продукт)",
    "Журнал (ППТС ID)",
    "Журнал (отв.)",
]


_MERGE_COLS = (1, 2, 3)  # columns merged per group: №, CVE ID, Продукт (ТСУ)


def _source_label(source: str) -> str:
    """Map candidate source to human-readable label."""
    if source == "local_ppts":
        return "ППТС лок."
    if source == "general_ppts":
        return "ППТС общ."
    if source == "knowledge_base":
        return "БЗ"
    return source  # For journal source_file names


def _write_detail_sheet(
    ws: Worksheet,
    results: list[AnalysisResult],
    settings: PipelineSettings | None = None,
) -> None:
    """Write the detailed analysis sheet (Sheet 2) with tier-filtered candidates."""
    ws.title = "Детальный анализ"

    for col, header in enumerate(_DETAIL_HEADERS, 1):
        ws.cell(row=1, column=col, value=header)
    _style_header(ws, len(_DETAIL_HEADERS))

    primary_limit = settings.detail_primary_limit if settings else 0
    secondary_limit = settings.detail_secondary_limit if settings else 3

    ncols = len(_DETAIL_HEADERS)
    row_idx = 2

    for result_num, r in enumerate(results, 1):
        tsu_display = _vendor_product(r.vulnerability.vendor, r.vulnerability.product)

        # Determine rows to write
        rows_to_write: list[tuple[str, MatchCandidate | None, JournalEntry | None]] = []

        # Add candidate rows
        filtered = (
            _filter_candidates(r.candidates, primary_limit, secondary_limit)
            if r.candidates
            else []
        )
        for c in filtered:
            rows_to_write.append(("candidate", c, None))

        # Add journal rows
        for jm in r.journal_matches:
            rows_to_write.append(("journal", None, jm))

        if not rows_to_write:
            continue

        group_start = row_idx

        for i, (row_type, candidate, journal_match) in enumerate(rows_to_write):
            # Shared columns: only on first row
            ws.cell(row=row_idx, column=1, value=result_num if i == 0 else None)
            ws.cell(row=row_idx, column=2, value=r.vulnerability.cve_id if i == 0 else None)
            ws.cell(row=row_idx, column=3, value=tsu_display if i == 0 else None)

            if row_type == "candidate":
                c = candidate  # type: ignore[assignment]
                ppts_display = _vendor_product(c.software.vendor, c.software.name)
                tier = _candidate_tier(c.combined_score)

                ws.cell(row=row_idx, column=4, value=ppts_display)
                ws.cell(row=row_idx, column=5, value=c.software.id)

                vs_cell = ws.cell(row=row_idx, column=6, value=round(c.vector_score, 4))
                ws.cell(row=row_idx, column=7, value=round(c.fuzzy_score, 2))
                ws.cell(row=row_idx, column=8, value=round(c.exact_score, 2))
                cs_cell = ws.cell(row=row_idx, column=9, value=round(c.combined_score, 4))

                for cell in (vs_cell, cs_cell):
                    fill = _score_fill(cell.value if cell.value else 0)
                    if fill:
                        cell.fill = fill

                ws.cell(row=row_idx, column=10, value=tier)
                ws.cell(row=row_idx, column=11, value=_source_label(c.software.source))
                # Journal columns empty for candidate rows

            elif row_type == "journal":
                jm = journal_match  # type: ignore[assignment]
                # Candidate columns 4-11 empty for journal rows
                ws.cell(row=row_idx, column=12, value=jm.source_file)
                product_j = jm.product if jm.product else ""
                ws.cell(row=row_idx, column=13, value=product_j)
                ws.cell(row=row_idx, column=14, value=jm.ppts_id or "")
                ws.cell(row=row_idx, column=15, value=jm.responsible or "")

            # Apply borders
            for col in range(1, ncols + 1):
                ws.cell(row=row_idx, column=col).border = _THIN_BORDER

            row_idx += 1

        group_end = row_idx - 1

        if group_end > group_start:
            for col in _MERGE_COLS:
                ws.merge_cells(
                    start_row=group_start, start_column=col,
                    end_row=group_end, end_column=col,
                )
                ws.cell(row=group_start, column=col).alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True,
                )

        for col in range(1, ncols + 1):
            cell = ws.cell(row=group_start, column=col)
            b = cell.border
            cell.border = Border(top=_SEP_SIDE, bottom=b.bottom, left=b.left, right=b.right)

    _auto_width(ws, ncols)


# ---------------------------------------------------------------------------
# Sheet 3: Reference
# ---------------------------------------------------------------------------


def _write_reference_sheet(ws: Worksheet, settings: PipelineSettings) -> None:
    """Write the reference/help sheet (Sheet 3)."""
    ws.title = "Справка"

    info = [
        ("Параметр", "Значение", "Описание"),
        ("Top-N", str(settings.top_n), "Количество кандидатов из векторного поиска"),
        ("Порог вектора", str(settings.vector_threshold), "Минимальный cosine similarity (0-1)"),
        ("Порог fuzzy", str(settings.fuzzy_threshold), "Минимальный fuzzy score (0-100)"),
        ("Транслитерация", settings.transliteration_direction, "Направление нормализации"),
        ("Мин. длина слова", str(settings.min_word_length), "Фильтр коротких слов для fuzzy"),
        ("База знаний", "Да" if settings.use_knowledge_base else "Нет", "Использование базы знаний"),
        ("", "", ""),
        ("Метрика", "Диапазон", "Описание"),
        ("Vector Score", "0.0 — 1.0", "Cosine similarity между эмбеддингами (sentence-transformers)"),
        ("Fuzzy Score", "0 — 100", "Лучший из token_sort_ratio, token_set_ratio, partial_ratio (RapidFuzz)"),
        ("Exact Score", "0 / 50 / 75 / 100", "100=точное совпадение, 75=подстрока, 50=все слова, 0=нет"),
        ("Итоговый балл", "0.0 — 1.0", "Взвешенная комбинация: 50% vector + 30% fuzzy + 20% exact"),
        ("", "", ""),
        ("Ранг", "Порог", "Описание"),
        ("1", "≥ 0.7", "Высокая уверенность"),
        ("2", "≥ 0.4", "Средняя уверенность"),
        ("3", "< 0.4", "Низкая уверенность"),
        ("", "", ""),
        ("Статус", "Источник", "Описание"),
        ("ДА", "Только БЗ", "ПО присутствует в инфраструктуре (назначается только через базу знаний)"),
        ("НЕТ", "Авто / БЗ", "ПО отсутствует в инфраструктуре"),
        ("ЛИНУКС", "БЗ", "Linux-специфичное ПО"),
        ("УСЛОВНО", "БЗ", "Требует дополнительной проверки"),
        ("(пусто)", "Авто", "Есть кандидаты, но уверенности недостаточно — требуется ручной разбор"),
    ]

    for row_idx, (a, b, c) in enumerate(info, 1):
        ws.cell(row=row_idx, column=1, value=a)
        ws.cell(row=row_idx, column=2, value=b)
        ws.cell(row=row_idx, column=3, value=c)

    for col in range(1, 4):
        cell = ws.cell(row=1, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT

    _auto_width(ws, 3, max_width=80)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def write_report(
    path: Path,
    results: list[AnalysisResult],
    settings: PipelineSettings,
    responsible: str = "",
    publication: str = "БДУ ФСТЕК",
) -> None:
    """Generate a full XLSX report with 3 sheets (main, detail, reference)."""
    wb = Workbook()

    ws_main = wb.active
    _write_main_sheet(ws_main, results, responsible, publication)

    ws_detail = wb.create_sheet()
    _write_detail_sheet(ws_detail, results, settings)

    ws_ref = wb.create_sheet()
    _write_reference_sheet(ws_ref, settings)

    wb.save(str(path))
    logger.info("Report saved to %s (%d results)", path, len(results))
